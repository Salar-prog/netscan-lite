import csv
import ipaddress
import logging
from pathlib import Path
from typing import List, Optional

from sqlmodel import Session, select

from netscan_lite.config import settings
from netscan_lite.models import Group, IPAddress, IPStatus

logger = logging.getLogger(__name__)


def validate_ip(ip_str: str) -> bool:
    """Validate an IPv4 address string."""
    try:
        ipaddress.IPv4Address(ip_str.strip())
        return True
    except ValueError:
        return False


def read_csv(file_path: Path) -> List[dict]:
    """Read a CSV file and return list of row dicts."""
    rows = []
    with open(file_path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append({k.strip().lower(): v for k, v in row.items() if k})
    return rows


def read_xlsx(file_path: Path) -> List[dict]:
    """Read an XLSX file and return list of row dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required for XLSX files. Install with: pip install openpyxl")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = str(value).strip() if value is not None else ""
        rows.append(row_dict)

    wb.close()
    return rows


def import_file(file_path: Path, session: Session, group_name: Optional[str] = None) -> dict:
    """Import IPs from CSV or XLSX file.

    Expected columns: ip (required), hostname (optional), group (optional)
    If group_name is provided, it overrides the 'group' column in the file.

    Returns dict with import stats.
    """
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        rows = read_csv(file_path)
    elif suffix == ".xlsx":
        rows = read_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file format: {suffix}. Use .csv or .xlsx")

    if not rows:
        return {"imported": 0, "skipped": 0, "errors": ["File is empty"]}

    stats = {"imported": 0, "skipped": 0, "errors": []}
    groups_cache: dict[str, Group] = {}

    for i, row in enumerate(rows, start=1):
        ip_str = row.get("ip", "").strip()
        if not ip_str:
            stats["skipped"] += 1
            stats["errors"].append(f"Row {i}: missing IP")
            continue

        if not validate_ip(ip_str):
            stats["skipped"] += 1
            stats["errors"].append(f"Row {i}: invalid IP '{ip_str}'")
            continue

        hostname = row.get("hostname", "").strip() or None
        ip_group = group_name or row.get("group", "").strip() or "default"

        if ip_group not in groups_cache:
            existing = session.exec(select(Group).where(Group.name == ip_group)).first()
            if existing:
                groups_cache[ip_group] = existing
            else:
                new_group = Group(
                    name=ip_group,
                    miss_threshold=settings.DEFAULT_MISS_THRESHOLD,
                    quarantine_hours=settings.DEFAULT_QUARANTINE_HOURS,
                )
                session.add(new_group)
                session.flush()
                groups_cache[ip_group] = new_group

        group = groups_cache[ip_group]

        existing_ip = session.exec(
            select(IPAddress).where(IPAddress.ip == ip_str, IPAddress.group_id == group.id)
        ).first()

        if existing_ip:
            stats["skipped"] += 1
            continue

        ip_obj = IPAddress(
            group_id=group.id,
            ip=ip_str,
            status=IPStatus.AVAILABLE_CANDIDATE,
            hostname=hostname,
        )
        session.add(ip_obj)
        stats["imported"] += 1

    session.commit()
    return stats
